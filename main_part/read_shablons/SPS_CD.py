import os
import shutil
from time import strftime
import time
import datetime

import openpyxl
import pandas as pd

def shablonExcel_SPS_CD(row, dataframes: list, organise_dct: dict, values_Excel):
    # Организационные моменты
    LAB_NO = organise_dct.get("LAB_NO")
    N_IG = organise_dct.get("N_IG")
    boreHole = organise_dct.get("boreHole")
    depth = organise_dct.get("depth")
    nameSoil = organise_dct.get("nameSoil")

    # Путь для сохранения протоколов
    pathSave = organise_dct.get("pathSave_unaxial")

    # Дата получение объекта подлежащего испытаниям
    date_isp_object = str(organise_dct.get("date_isp_object"))

    # Дата испытания
    date_isp = str(organise_dct.get("date_isp"))

    # Дата протокола
    date_protocol = str(organise_dct.get("date_protocol"))

    # Номер протокола
    number_protocol = organise_dct.get("number_protocol")

    # Заказчик
    nameClient = organise_dct.get("nameClient")

    # Объект
    nameObject = organise_dct.get("nameObject")

    # Параметры образца
    We = organise_dct.get('We')
    p = organise_dct.get('p')
    ps = organise_dct.get('ps')
    e = organise_dct.get('e')
    IL = organise_dct.get('IL')

    if str(LAB_NO) in ["NAN", "None", "nAn"]:
        prot_name = str(row) + '.xlsx'
    else:
        prot_name = LAB_NO + '.xlsx'

    """
    Бот '..\\srcs\\shablons\\TPDS_CD_test.xlsx'
    Локальная '..\\GEOF\\srcs\\shablons\\TPDS_CD_test.xlsx'
    """
    shutil.copy('..\\GEOF\\srcs\\shablons\\SPS_CD.xlsx'
                ,f'{pathSave}\\{prot_name}')


    wb = openpyxl.load_workbook(
        f'{pathSave}\\{prot_name}')
    ws = wb.active

    ws['O20'] = LAB_NO
    ws['O21'] = boreHole
    ws['O22'] = depth
    ws['O23'] = nameSoil

    ws['U20'] = We
    ws['U21'] = p
    ws['U22'] = ps
    ws['U23'] = e
    ws['U24'] = IL

    ### Модули
    ws['A65'] = values_Excel.get('devE50')
    ws['B65'] = values_Excel.get('epsE50')
    ws['B70'] = values_Excel.get('devE0')
    ws['A70'] = values_Excel.get('epsE0')

    ws['O53'] = organise_dct.get('F_unaxial')
    ws['O54'] = organise_dct.get('C_unaxial')

    # Давления
    ws['N47'] = organise_dct.get('Start1_unaxial')
    ws['N48'] = organise_dct.get('Start2_unaxial')
    ws['N49'] = organise_dct.get('Start3_unaxial')

    ws['O47'] = organise_dct.get('End1_unaxial')
    ws['O48'] = organise_dct.get('End2_unaxial')
    ws['O49'] = organise_dct.get('End3_unaxial')

    date_protocol = [str(x) for x in date_protocol.replace(' 00:00:00', '').split('-')]
    date_protocol.reverse()
    str_prot = "Протокол испытаний № " + number_protocol + ' от ' + str('-'.join(map(str, date_protocol)))
    ws['M9'] = str_prot

    ws['M11'] = 'Наименование и адрес заказчика: ' + nameClient
    ws['M12'] = 'Наименование объекта: ' + nameObject

    date_isp_object = [str(x) for x in date_isp_object.replace(' 00:00:00', '').split('-')]
    date_isp_object.reverse()
    ws['M15'] = 'Дата получение объекта подлежащего испытаниям: ' + str('-'.join(map(str, date_isp_object)))

    if isinstance(date_isp, str):
        ws['M16'] = 'Дата испытания: ' + date_isp
    else:
        ws['M16'] = 'Дата испытания: ' + date_isp

    wb.save(f'{pathSave}\\{prot_name}')

    writer = pd.ExcelWriter(
        f'{pathSave}\\{prot_name}', mode='a',
        engine="openpyxl",
        if_sheet_exists='overlay')

    dataframe1 = dataframes[0]
    dataframe2 = dataframes[1]
    dataframe3 = dataframes[2]

    dataframe1.to_excel(writer, sheet_name='1', startcol=5, startrow=64, index=False,
                        index_label=False,
                        header=False)

    dataframe1.to_excel(writer, sheet_name='1', startcol=9, startrow=64, index=False,
                        index_label=False,
                        header=False)

    dataframe2.to_excel(writer, sheet_name='1', startcol=11, startrow=64, index=False,
                        index_label=False,
                        header=False)

    dataframe3.to_excel(writer, sheet_name='1', startcol=13, startrow=64, index=False,
                        index_label=False,
                        header=False)

    writer.close()