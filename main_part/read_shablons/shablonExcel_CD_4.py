import math
import os
import shutil
from time import strftime
import time
import datetime

import openpyxl
import pandas as pd

def shablonExcel_TPS_CD_4(row, dataframes: list, organise_dct: dict, values_Excel: dict, mode: int) -> None:
    # Организационные моменты
    LAB_NO = organise_dct.get("LAB_NO")
    N_IG = organise_dct.get("N_IG")
    boreHole = organise_dct.get("boreHole")
    depth = organise_dct.get("depth")
    nameSoil = organise_dct.get("nameSoil")


    # Путь для сохранения протоколов
    pathSave = organise_dct.get("pathSave_traxial_CD")

    print(pathSave)


    # Дата получение объекта подлежащего испытаниям
    date_isp_object = str(organise_dct.get("date_isp_object"))

    # Дата испытания
    date_isp = str(organise_dct.get("date_isp"))

    # Дата протокола
    date_protocol = str(organise_dct.get("date_protocol"))

    # Номер протокола
    number_protocol = organise_dct.get("number_protocol")

    # Заказчик
    nameClient = organise_dct.get("nameClient")

    # Объект
    nameObject = organise_dct.get("nameObject")

    # Параметры образца
    We = organise_dct.get('We')
    p = organise_dct.get('p')
    ps = organise_dct.get('ps')
    e = organise_dct.get('e')
    IL = organise_dct.get('IL')

    if str(LAB_NO) in ["NAN", "None", "nAn"]:
        prot_name = str(row) + '.xlsx'
    else:
        prot_name = LAB_NO + '.xlsx'

    """
    Бот '..\\srcs\\shablons\\TPDS_CD_test.xlsx'
    Локальная '..\\GEOF\\srcs\\shablons\\TPDS_CD_test.xlsx'
    """
    shutil.copy('..\\GEOF\\srcs\\shablons\\TPDS_CD_test.xlsx'
                , f'{pathSave}\\{prot_name}')

    wb = openpyxl.load_workbook(
        f'{pathSave}\\{prot_name}')
    ws = wb.active

    ws['C17'] = LAB_NO
    ws['C18'] = boreHole
    ws['C19'] = depth
    ws['C20'] = nameSoil

    ws['J17'] = We
    ws['J18'] = p
    ws['J19'] = ps
    ws['J20'] = e
    ws['J21'] = IL

    ### Модули
    ws['D85'] = values_Excel.get('devE50')
    ws['E85'] = values_Excel.get('epsE50')

    ws['A85'] = values_Excel.get('devE0')
    ws['B85'] = values_Excel.get('epsE0')

    ws['O56'] = organise_dct.get('F_traxial')
    ws['O57'] = organise_dct.get('C_traxial')

    # Давления
    # K0, д.е.
    ws['D62'] = (1 - math.sin(math.radians(organise_dct.get('F_traxial'))))
    # Эффективное напряжение, Мпа:
    ws['J62'] = organise_dct.get('pressStart1_traxial') * (1 - math.sin(math.radians(organise_dct.get('F_traxial'))))
    # Точки нахождения модуля Е0, Мпа (полное напряжение):
    # Первая точка модуля E0
    ws['J63'] = organise_dct.get('pressStart1_traxial') * (1 - math.sin(math.radians(organise_dct.get('F_traxial'))))
    # Вторая точка модуля E0
    ws['K63'] = organise_dct.get('pressStart1_traxial') * (1 - math.sin(math.radians(organise_dct.get('F_traxial')))) + values_Excel.get('devE0')
    # Максимальный девиатор
    ws['J64'] = values_Excel.get('devMAX')
    # Точка на Е50
    ws['J65'] = values_Excel.get('devE50')

    ws['N47'] = organise_dct.get('pressStart1_traxial')
    ws['N48'] = organise_dct.get('pressStart2_traxial')
    ws['N49'] = organise_dct.get('pressStart3_traxial')

    ws['O47'] = organise_dct.get('pressEnd1_traxial')
    ws['O48'] = organise_dct.get('pressEnd2_traxial')
    ws['O49'] = organise_dct.get('pressEnd3_traxial')

    date_protocol = [str(x) for x in date_protocol.replace(' 00:00:00', '').split('-')]
    date_protocol.reverse()
    str_prot = "Протокол испытаний № " + number_protocol + ' от ' + str('-'.join(map(str, date_protocol)))
    ws['A9'] = str_prot

    ws['A10'] = 'Наименование и адрес заказчика: ' + nameClient
    ws['A11'] = 'Наименование объекта: ' + nameObject

    date_isp_object = [str(x) for x in date_isp_object.replace(' 00:00:00', '').split('-')]
    date_isp_object.reverse()
    ws['A14'] = 'Дата получение объекта подлежащего испытаниям: ' + str('-'.join(map(str, date_isp_object)))

    if isinstance(date_isp, str):
        ws['A15'] = 'Дата испытания: ' + date_isp
    else:
        ws['A15'] = 'Дата испытания: ' + date_isp

    wb.save(f'{pathSave}\\{prot_name}')

    writer = pd.ExcelWriter(
        f'{pathSave}\\{prot_name}', mode='a',
        engine="openpyxl",
        if_sheet_exists='overlay')

    if mode in [2, 3]:
        dataframe0 = dataframes[0]
        dataframe1 = dataframes[1]
        dataframe2 = dataframes[2]
        dataframe3 = dataframes[3]

        dataframe0.to_excel(writer, sheet_name='1', startcol=5, startrow=84, index=False,
                            index_label=False,
                            header=False)

        dataframe1.to_excel(writer, sheet_name='1', startcol=9, startrow=84, index=False,
                            index_label=False,
                            header=False)

        dataframe2.to_excel(writer, sheet_name='1', startcol=12, startrow=84, index=False,
                            index_label=False,
                            header=False)

        dataframe3.to_excel(writer, sheet_name='1', startcol=15, startrow=84, index=False,
                            index_label=False,
                            header=False)

    else:
        dataframe0 = dataframes[0]
        dataframe1 = dataframes[1]
        dataframe2 = dataframes[2]

        dataframe0.to_excel(writer, sheet_name='1', startcol=5, startrow=84, index=False,
                            index_label=False,
                            header=False)

        dataframe0.to_excel(writer, sheet_name='1', startcol=9, startrow=84, index=False,
                            index_label=False,
                            header=False)

        dataframe1.to_excel(writer, sheet_name='1', startcol=12, startrow=84, index=False,
                            index_label=False,
                            header=False)

        dataframe2.to_excel(writer, sheet_name='1', startcol=15, startrow=84, index=False,
                            index_label=False,
                            header=False)
    writer.close()