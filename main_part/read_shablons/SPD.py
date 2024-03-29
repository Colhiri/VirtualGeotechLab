import shutil
import io

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image

def shablonExcel_SPD(row, dataframes: list, organise_dct: dict, values_Excel):
    # Организационные моменты
    LAB_NO = organise_dct.get("LAB_NO")
    N_IG = organise_dct.get("N_IG")
    boreHole = organise_dct.get("boreHole")
    depth = organise_dct.get("depth")
    nameSoil = organise_dct.get("nameSoil")

    # Путь для сохранения протоколов
    pathSave = organise_dct.get("pathSave_spd")

    # Дата получение объекта подлежащего испытаниям
    date_isp_object = str(organise_dct.get("date_isp_object"))

    # Дата испытания
    date_isp = str(organise_dct.get("date_isp"))

    # Дата протокола
    date_protocol = str(organise_dct.get("date_protocol"))

    # Номер протокола
    number_protocol = organise_dct.get("number_protocol")

    # Заказчик
    nameClient = organise_dct.get("nameClient")

    # Объект
    nameObject = organise_dct.get("nameObject")

    # Параметры образца
    We = organise_dct.get('We')
    p = organise_dct.get('p')
    ps = organise_dct.get('ps')
    e = organise_dct.get('e')
    IL = organise_dct.get('IL')

    if str(LAB_NO) in ['None', 'nan', 'nAn', 'NA', '<NA>']:
        prot_name = str(row) + '.xlsx'
    else:
        prot_name = LAB_NO + '.xlsx'

    """
    Бот '..\\srcs\\shablons\\TPDS_CD_test.xlsx'
    Локальная '..\\GEOF\\srcs\\shablons\\TPDS_CD_test.xlsx'
    """
    shutil.copy('..\\GEOF\\srcs\\shablons\\SPD.xlsx'
                    ,f'{pathSave}\\{prot_name}')

    wb = openpyxl.load_workbook(
        f'{pathSave}\\{prot_name}')
    ws = wb.active

    ws['C20'] = LAB_NO
    ws['C21'] = boreHole
    ws['C22'] = depth
    ws['C23'] = nameSoil

    ws['I20'] = We
    ws['I21'] = p
    ws['I22'] = ps
    ws['I23'] = e
    ws['I24'] = IL

    ### Значения для касательного модуля
    ws['G36'] = values_Excel.get('q_zg')
    ws['G37'] = values_Excel.get('otn_zg')
    ws['I36'] = values_Excel.get('otn_END_A')
    ws['I37'] = values_Excel.get('otn_MAX')
    ws['K37'] = values_Excel.get('press_MAX')

    date_protocol = [str(x) for x in date_protocol.replace(' 00:00:00', '').split('-')]
    date_protocol.reverse()
    str_prot = "Протокол испытаний № " + number_protocol + ' от ' + str('-'.join(map(str, date_protocol)))
    ws['A9'] = str_prot

    ws['A11'] = 'Наименование и адрес заказчика: ' + nameClient
    ws['A12'] = 'Наименование объекта: ' + nameObject

    date_isp_object = [str(x) for x in date_isp_object.replace(' 00:00:00', '').split('-')]
    date_isp_object.reverse()
    ws['A15'] = 'Дата получение объекта подлежащего испытаниям: ' + str('-'.join(map(str, date_isp_object)))

    if isinstance(date_isp, str):
        ws['A16'] = 'Дата испытания: ' + date_isp
    else:
        ws['A16'] = 'Дата испытания: ' + date_isp

    # ws.add_image(image_graph, anchor='E29')


    ws['A55'] = f'Инженерно-геологический элемент: {N_IG}'

    wb.save(f'{pathSave}\\{prot_name}')

    writer = pd.ExcelWriter(
        f'{pathSave}\\{prot_name}', mode='a',
        engine="openpyxl",
        if_sheet_exists='overlay')

    dataframe1 = dataframes[0]

    dataframe1 = dataframe1.astype('float64', errors='ignore')

    dataframe1.to_excel(writer, sheet_name='1', startcol=5, startrow=29, index=False,
                        index_label=False,
                        header=False, float_format="%.20f")

    writer.close()
