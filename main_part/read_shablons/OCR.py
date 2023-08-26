import math
import random
import os
import shutil
from time import strftime
import time
import datetime

import numpy as np
import openpyxl
import pandas as pd

def shablonExcel_OCR(row, dataframes: list, organise_dct: dict, values_Excel):
    # Организационные моменты
    LAB_NO = organise_dct.get("LAB_NO")
    N_IG = organise_dct.get("N_IG")
    boreHole = organise_dct.get("boreHole")
    depth = organise_dct.get("depth")
    nameSoil = organise_dct.get("nameSoil")

    # Путь для сохранения протоколов
    pathSave = organise_dct.get("pathSave_OCR")

    # Дата получение объекта подлежащего испытаниям
    date_isp_object = str(organise_dct.get("date_isp_object"))

    # Дата испытания
    date_isp = str(organise_dct.get("date_isp"))

    # Дата протокола
    date_protocol = str(organise_dct.get("date_protocol"))

    # Номер протокола
    number_protocol = organise_dct.get("number_protocol")

    # Заказчик
    nameClient = organise_dct.get("nameClient")

    # Объект
    nameObject = organise_dct.get("nameObject")

    # Параметры образца
    We = organise_dct.get('We')
    p = organise_dct.get('p')
    ps = organise_dct.get('ps')
    e = organise_dct.get('e')
    IL = organise_dct.get('IL')

    if str(LAB_NO) in ['None', 'nan', 'nAn', 'NA', '<NA>']:
        prot_name = str(row) + '.xlsx'
    else:
        prot_name = LAB_NO + '.xlsx'

    shutil.copy('..\\GEOF\\srcs\\shablons\\OCR.xlsx'
                , f'{pathSave}\\{prot_name}')

    wb = openpyxl.load_workbook(
        f'{pathSave}\\{prot_name}')
    ws = wb.active

    ws['C20'] = LAB_NO
    ws['C21'] = boreHole
    ws['C22'] = depth
    ws['C23'] = nameSoil

    ws['I20'] = We
    ws['I21'] = p
    ws['I22'] = ps
    ws['I23'] = e
    ws['I24'] = IL

    date_protocol = [str(x) for x in date_protocol.replace(' 00:00:00', '').split('-')]
    date_protocol.reverse()
    str_prot = "Протокол испытаний № " + number_protocol + ' от ' + str('-'.join(map(str, date_protocol)))
    ws['A9'] = str_prot

    ws['A11'] = 'Наименование и адрес заказчика: ' + nameClient
    ws['A12'] = 'Наименование объекта: ' + nameObject

    date_isp_object = [str(x) for x in date_isp_object.replace(' 00:00:00', '').split('-')]
    date_isp_object.reverse()
    ws['A15'] = 'Дата получение объекта подлежащего испытаниям: ' + str('-'.join(map(str, date_isp_object)))

    if isinstance(date_isp, str):
        ws['A16'] = 'Дата испытания: ' + date_isp
    else:
        ws['A16'] = 'Дата испытания: ' + date_isp

    # Казагранде
    ws['K38'] = values_Excel.get('Sigma_CASAGRANDE')
    ws['K39'] = values_Excel.get('effective_press')
    ws['K40'] = values_Excel.get('OCR_CASAGRANDE')
    ws['K41'] = values_Excel.get('POP_CASAGRANDE')

    # Беккер
    ws['K55'] = values_Excel.get('Sigma_Beccer')
    ws['K56'] = values_Excel.get('effective_press')
    ws['K57'] = values_Excel.get('OCR')
    ws['K58'] = values_Excel.get('POP')

    ws['A67'] = f'Инженерно-геологический элемент: {N_IG}'

    # Изображения
    # ws.add_image(values_Excel.get('image_casagrande'), 'F26')
    # ws.add_image(values_Excel.get('image_beccer'), 'F43')

    wb.save(f'{pathSave}\\{prot_name}')

    writer = pd.ExcelWriter(
        f'{pathSave}\\{prot_name}', mode='a',
        engine="openpyxl",
        if_sheet_exists='overlay')

    dataframe1 = dataframes[0]
    dataframe1 = dataframe1.astype('float64')

    dataframe2 = dataframes[1]
    dataframe2 = dataframe2.astype('float64')

    # метод Казагранде
    (dataframe1).to_excel(writer, sheet_name='1', startcol=1, startrow=26, index=False,
                          index_label=False,
                          header=False, float_format="%.20f")
    # Значения для линий
    CASAGRANDE = values_Excel.get('VALUES_LINES').astype('float64')
    CASAGRANDE.to_excel(writer, sheet_name='1', startcol=13, startrow=26, index=False,
                        index_label=False,
                        header=False, float_format="%.20f")


    # метод Беккера
    (dataframe2).to_excel(writer, sheet_name='1', startcol=1, startrow=43, index=False,
                        index_label=False,
                        header=False, float_format="%.20f")
    # Значения для линий
    BECCER = values_Excel.get('BECCER').astype('float64')
    BECCER.to_excel(writer, sheet_name='1', startcol=13, startrow=43, index=False,
                        index_label=False,
                        header=False, float_format="%.20f")

    writer.close()